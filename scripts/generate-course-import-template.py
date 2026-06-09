#!/usr/bin/env python3
"""Generate course import workbook with dropdown validation."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

HEADERS = [
    "Sector Name",
    "Linked Program",
    "Linked Scheme",
    "Course Name",
    "SIDH Course ID",
    "Job Role",
    "NSQF Level",
    "Training Per Day (Hours)",
    "Approval Status",
    "Approval Date",
    "Total Hours",
    "Valid Until",
    "Short Form",
]

DEFAULT_SECTOR_NAMES: list[str] = []
DEFAULT_PROGRAM_NAMES: list[str] = []
DEFAULT_SCHEME_NAMES: list[str] = []
DEFAULT_APPROVAL_STATUS_OPTIONS = ["Pending", "Approved"]

DEFAULT_OUTPUT_PATH = Path(__file__).resolve().parents[1] / "public" / "course_import_template.xlsx"
MAX_DATA_ROW = 50_001


def write_list_column(sheet, column: int, values: list[str]) -> str | None:
    if not values:
        return None

    for index, value in enumerate(values, start=1):
        sheet.cell(row=index, column=column, value=value)

    column_letter = sheet.cell(row=1, column=column).column_letter
    return f"Lists!${column_letter}$1:${column_letter}${len(values)}"


def add_list_validation(sheet, cell_range: str, formula: str, *, allow_blank: bool) -> None:
    validation = DataValidation(
        type="list",
        formula1=formula if formula.startswith("=") else f"={formula}",
        allow_blank=allow_blank,
        showDropDown=False,
    )
    sheet.add_data_validation(validation)
    validation.add(cell_range)


def build_workbook(
    sector_names: list[str],
    program_names: list[str],
    scheme_names: list[str],
    approval_status_options: list[str],
) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Course Import Template"

    lists_sheet = workbook.create_sheet("Lists")
    lists_sheet.sheet_state = "hidden"

    sector_source = write_list_column(lists_sheet, 1, sector_names)
    program_source = write_list_column(lists_sheet, 2, program_names)
    scheme_source = write_list_column(lists_sheet, 3, scheme_names)
    approval_source = write_list_column(lists_sheet, 4, approval_status_options)

    sample_sector = sector_names[0] if sector_names else "Agriculture"
    sample_program = program_names[0] if program_names else "Skill Development Program"
    sample_scheme = scheme_names[0] if scheme_names else "PMKVY"

    sample_row = [
        sample_sector,
        sample_program,
        sample_scheme,
        "Maize Cultivator",
        "FeeSchCor_48128",
        "Kisan Drone Operator",
        "4",
        6,
        "Approved",
        "01/01/2026",
        320,
        "31/12/2028",
        "MC",
    ]

    sheet.append(HEADERS)
    sheet.append(sample_row)

    if sector_source:
        add_list_validation(sheet, f"A2:A{MAX_DATA_ROW}", sector_source, allow_blank=False)
    if program_source:
        add_list_validation(sheet, f"B2:B{MAX_DATA_ROW}", program_source, allow_blank=False)
    if scheme_source:
        add_list_validation(sheet, f"C2:C{MAX_DATA_ROW}", scheme_source, allow_blank=False)
    if approval_source:
        add_list_validation(sheet, f"I2:I{MAX_DATA_ROW}", approval_source, allow_blank=False)

    return workbook


def load_config(config_path: Path | None) -> tuple[list[str], list[str], list[str], list[str]]:
    if config_path is None:
        return DEFAULT_SECTOR_NAMES, DEFAULT_PROGRAM_NAMES, DEFAULT_SCHEME_NAMES, DEFAULT_APPROVAL_STATUS_OPTIONS

    payload = json.loads(config_path.read_text(encoding="utf-8"))
    sector_names = [str(value).strip() for value in payload.get("sectorNames", []) if str(value).strip()]
    program_names = [str(value).strip() for value in payload.get("programNames", []) if str(value).strip()]
    scheme_names = [str(value).strip() for value in payload.get("schemeNames", []) if str(value).strip()]
    approval_status_options = [
        str(value).strip()
        for value in payload.get("approvalStatusOptions", DEFAULT_APPROVAL_STATUS_OPTIONS)
        if str(value).strip()
    ]
    return sector_names, program_names, scheme_names, approval_status_options


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", type=Path, default=None)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_PATH)
    args = parser.parse_args()

    sector_names, program_names, scheme_names, approval_status_options = load_config(args.config)
    workbook = build_workbook(sector_names, program_names, scheme_names, approval_status_options)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(f"Wrote {args.output}")


if __name__ == "__main__":
    main()
