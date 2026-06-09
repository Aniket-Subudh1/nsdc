#!/usr/bin/env python3
"""Generate candidate import workbook with dropdown validation."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

NAME_PREFIX_OPTIONS = ["Mr", "Mrs", "Ms", "Mx"]
GENDER_OPTIONS = ["Male", "Female", "Transgender"]

HEADERS = [
    "Name Prefix",
    "Full Name",
    "Gender",
    "DOB",
    "Father's Name",
    "Guardian Name",
    "Email",
    "Phone",
    "Country Code",
    "State",
    "District",
    "Center Name",
    "Course (reference only)",
]

DEFAULT_CENTER_NAMES: list[str] = []
DEFAULT_COURSE_NAMES: list[str] = []

DEFAULT_OUTPUT_PATH = Path(__file__).resolve().parents[1] / "public" / "candidate_details.xlsx"
DEFAULT_LOCATION_OPTIONS_PATH = Path(__file__).resolve().parents[1] / "lib" / "candidate-location-options.json"
MAX_DATA_ROW = 50_001
FIRST_STATE_CITY_COLUMN = 6


def write_list_column(sheet, column: int, values: list[str]) -> str | None:
    if not values:
        return None

    for index, value in enumerate(values, start=1):
        sheet.cell(row=index, column=column, value=value)

    column_letter = sheet.cell(row=1, column=column).column_letter
    return f"Lists!${column_letter}$1:${column_letter}${len(values)}"


def add_list_validation(sheet, cell_range: str, formula: str, *, allow_blank: bool) -> None:
    validation = DataValidation(
        type="list",
        formula1=formula if formula.startswith("=") else f"={formula}",
        allow_blank=allow_blank,
        # openpyxl inverts this flag: False shows the Excel dropdown arrow.
        showDropDown=False,
    )
    sheet.add_data_validation(validation)
    validation.add(cell_range)


def load_location_options(path: Path) -> dict[str, list[str]]:
    if not path.exists():
        return {}

    payload = json.loads(path.read_text(encoding="utf-8"))
    return {
        str(state).strip(): [str(city).strip() for city in cities if str(city).strip()]
        for state, cities in payload.items()
        if str(state).strip()
    }


def excel_defined_name(state: str) -> str:
    sanitized = re.sub(r"[^A-Za-z0-9_.]", "_", state.strip())
    sanitized = re.sub(r"_+", "_", sanitized).strip("_")
    if sanitized and sanitized[0].isdigit():
        sanitized = f"_{sanitized}"
    return sanitized or "STATE"


def build_workbook(
    center_names: list[str],
    course_names: list[str],
    state_cities: dict[str, list[str]],
) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Candidates"

    lists_sheet = workbook.create_sheet("Lists")
    lists_sheet.sheet_state = "hidden"

    prefix_source = write_list_column(lists_sheet, 1, NAME_PREFIX_OPTIONS)
    gender_source = write_list_column(lists_sheet, 2, GENDER_OPTIONS)
    center_source = write_list_column(lists_sheet, 3, center_names)
    course_source = write_list_column(lists_sheet, 4, course_names)

    states = list(state_cities.keys())
    state_source = write_list_column(lists_sheet, 5, states)

    used_names: set[str] = set()
    for index, state in enumerate(states):
        cities = state_cities[state]
        if not cities:
            continue

        column = FIRST_STATE_CITY_COLUMN + index
        city_source = write_list_column(lists_sheet, column, cities)
        if not city_source:
            continue

        defined_name = excel_defined_name(state)
        suffix = 1
        while defined_name in used_names:
            suffix += 1
            defined_name = f"{excel_defined_name(state)}_{suffix}"
        used_names.add(defined_name)

        workbook.defined_names.add(DefinedName(defined_name, attr_text=city_source.replace("=", "")))

    sample_state = "ODISHA" if "ODISHA" in state_cities else (states[0] if states else "ODISHA")
    sample_district = state_cities.get(sample_state, ["CUTTACK"])[0] if state_cities.get(sample_state) else "CUTTACK"
    sample_center = center_names[0] if center_names else "Center One"
    sample_course = course_names[0] if course_names else None

    sample_row = [
        "Mr",
        "Rohit Kumar",
        "Male",
        "10/06/2005",
        "Suresh Kumar",
        None,
        "rohit@example.com",
        "9876543210",
        "91",
        sample_state,
        sample_district,
        sample_center,
        sample_course,
    ]

    sheet.append(HEADERS)
    sheet.append(sample_row)

    add_list_validation(sheet, f"A2:A{MAX_DATA_ROW}", prefix_source or '""', allow_blank=False)
    add_list_validation(sheet, f"C2:C{MAX_DATA_ROW}", gender_source or '""', allow_blank=False)

    if state_source:
        add_list_validation(sheet, f"J2:J{MAX_DATA_ROW}", state_source, allow_blank=True)

    if states:
        add_list_validation(
            sheet,
            f"K2:K{MAX_DATA_ROW}",
            'INDIRECT(SUBSTITUTE($J2," ","_"))',
            allow_blank=True,
        )

    if center_source:
        add_list_validation(sheet, f"L2:L{MAX_DATA_ROW}", center_source, allow_blank=False)

    if course_source:
        add_list_validation(sheet, f"M2:M{MAX_DATA_ROW}", course_source, allow_blank=True)

    return workbook


def load_config(config_path: Path | None) -> tuple[list[str], list[str]]:
    if config_path is None:
        return DEFAULT_CENTER_NAMES, DEFAULT_COURSE_NAMES

    payload = json.loads(config_path.read_text(encoding="utf-8"))
    center_names = [str(value).strip() for value in payload.get("centerNames", []) if str(value).strip()]
    course_names = [str(value).strip() for value in payload.get("courseNames", []) if str(value).strip()]
    return center_names, course_names


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", type=Path, default=None)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_PATH)
    parser.add_argument("--location-options", type=Path, default=DEFAULT_LOCATION_OPTIONS_PATH)
    args = parser.parse_args()

    center_names, course_names = load_config(args.config)
    state_cities = load_location_options(args.location_options)
    workbook = build_workbook(center_names, course_names, state_cities)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(f"Wrote {args.output}")


if __name__ == "__main__":
    main()
