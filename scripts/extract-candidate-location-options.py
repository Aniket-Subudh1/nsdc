#!/usr/bin/env python3
"""Extract SIDH LGD state/city options from the bulk candidate upload template."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_SOURCE = Path.home() / "Downloads/SIDH_API Integration @GTET 2/SIDH Bulk Candidate_upload Template.xlsx"
DEFAULT_OUTPUT = Path(__file__).resolve().parents[1] / "lib/candidate-location-options.ts"
DEFAULT_JSON_OUTPUT = Path(__file__).resolve().parents[1] / "lib/candidate-location-options.json"

NON_STATE_HEADERS = {
    "Type Of ID",
    "yesno",
    "EmploymentStatus",
    "HeardAboutUs",
    "Education List",
    "TrainingStatus",
    "State",
    "AadharID",
}


def clean(value: object | None) -> str | None:
    if value is None:
        return None

    text = str(value).replace("\ufeff", "").strip()
    if not text or text.lower() == "none":
        return None

    return " ".join(text.split())


def normalize_state_key(value: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", value.upper())


def read_state_labels(sheet) -> list[str]:
    headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    state_column_index = next(index for index, header in enumerate(headers) if header == "State")

    labels: list[str] = []
    seen: set[str] = set()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if state_column_index >= len(row):
            continue

        label = clean(row[state_column_index])
        if not label:
            continue

        key = normalize_state_key(label)
        if key in seen:
            continue

        seen.add(key)
        labels.append(label)

    return labels


def extract_state_city_map(workbook_path: Path) -> dict[str, list[str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook["Master Reference Data"]
    headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))

    state_labels = read_state_labels(sheet)
    label_by_key = {normalize_state_key(label): label for label in state_labels}
    state_cities = {label: [] for label in state_labels}

    for column_index, header in enumerate(headers):
        if column_index < 10 or not header:
            continue

        header_name = str(header).strip()
        if header_name in NON_STATE_HEADERS or "Constituency" in header_name:
            continue

        state_label = label_by_key.get(normalize_state_key(header_name))
        if not state_label:
            continue

        cities: list[str] = []
        seen: set[str] = set()

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if column_index >= len(row):
                continue

            city = clean(row[column_index])
            if not city:
                continue

            key = city.upper()
            if key in seen:
                continue

            seen.add(key)
            cities.append(city)

        state_cities[state_label] = cities

    return {label: state_cities[label] for label in state_labels if state_cities.get(label)}


def render_typescript(state_cities: dict[str, list[str]]) -> str:
    states = list(state_cities.keys())

    return f"""// Auto-generated from SIDH Bulk Candidate_upload Template.xlsx (Master Reference Data).
// Regenerate with: python3 scripts/extract-candidate-location-options.py

export const CANDIDATE_STATE_OPTIONS = {json.dumps(states, indent=2)} as const;

export type CandidateState = (typeof CANDIDATE_STATE_OPTIONS)[number];

export const CANDIDATE_STATE_DISTRICT_MAP: Record<CandidateState, readonly string[]> = {json.dumps(state_cities, indent=2)} as const;

function normalizeLocationToken(value: string) {{
  return value.trim().replace(/\\s+/g, " ");
}}

function normalizeStateLookupKey(value: string) {{
  return normalizeLocationToken(value).toUpperCase().replace(/[^A-Z0-9]/g, "");
}}

function normalizeDistrictLookupKey(value: string) {{
  return normalizeLocationToken(value).toUpperCase().replace(/\\./g, "").replace(/[^A-Z0-9]/g, "");
}}

const CANDIDATE_STATE_LOOKUP = Object.fromEntries(
  CANDIDATE_STATE_OPTIONS.map((state) => [normalizeStateLookupKey(state), state]),
) as Record<string, CandidateState>;

const CANDIDATE_DISTRICT_LOOKUP = Object.fromEntries(
  CANDIDATE_STATE_OPTIONS.map((state) => [
    state,
    Object.fromEntries(
      CANDIDATE_STATE_DISTRICT_MAP[state].map((district) => [normalizeDistrictLookupKey(district), district]),
    ),
  ]),
) as Record<CandidateState, Record<string, string>>;

export function resolveCandidateState(value: string) {{
  const trimmed = normalizeLocationToken(value);
  if (!trimmed) {{
    return "";
  }}

  const exactMatch = CANDIDATE_STATE_OPTIONS.find((option) => option.toLowerCase() === trimmed.toLowerCase());
  if (exactMatch) {{
    return exactMatch;
  }}

  return CANDIDATE_STATE_LOOKUP[normalizeStateLookupKey(trimmed)] ?? "";
}}

export function resolveCandidateDistrict(state: string, value: string) {{
  const trimmed = normalizeLocationToken(value);
  if (!trimmed) {{
    return "";
  }}

  const resolvedState = resolveCandidateState(state);
  if (!resolvedState) {{
    return "";
  }}

  const districts = CANDIDATE_STATE_DISTRICT_MAP[resolvedState];
  const exactMatch = districts.find((option) => option.toLowerCase() === trimmed.toLowerCase());
  if (exactMatch) {{
    return exactMatch;
  }}

  return CANDIDATE_DISTRICT_LOOKUP[resolvedState][normalizeDistrictLookupKey(trimmed)] ?? "";
}}

export function listCandidateDistrictsForState(state: string) {{
  const resolvedState = resolveCandidateState(String(state ?? ""));
  if (!resolvedState) {{
    return [] as string[];
  }}

  return [...CANDIDATE_STATE_DISTRICT_MAP[resolvedState]];
}}

export function normalizeCandidateState(value: unknown) {{
  return resolveCandidateState(String(value ?? ""));
}}

export function normalizeCandidateDistrict(state: unknown, value: unknown) {{
  return resolveCandidateDistrict(String(state ?? ""), String(value ?? ""));
}}

export function isKnownCandidateState(value: unknown) {{
  const trimmed = normalizeLocationToken(String(value ?? ""));
  if (!trimmed) {{
    return true;
  }}

  return resolveCandidateState(trimmed) !== "";
}}

export function isKnownCandidateDistrictForState(state: unknown, value: unknown) {{
  const trimmed = normalizeLocationToken(String(value ?? ""));
  if (!trimmed) {{
    return true;
  }}

  return resolveCandidateDistrict(String(state ?? ""), trimmed) !== "";
}}

export const CANDIDATE_STATE_ERROR = `State must be one of the SIDH LGD values: ${{CANDIDATE_STATE_OPTIONS.slice(0, 5).join(", ")}}, ... (${{CANDIDATE_STATE_OPTIONS.length}} total)`;

export function candidateDistrictError(state: string) {{
  const districts = listCandidateDistrictsForState(state);
  const resolvedState = normalizeCandidateState(state);
  if (districts.length === 0) {{
    return "District must match a SIDH LGD value for the selected state";
  }}

  return `District must be one of the SIDH LGD values for ${{resolvedState || state}}: ${{districts.slice(0, 5).join(", ")}}${{districts.length > 5 ? ", ..." : ""}}`;
}}

/** @deprecated Use listCandidateDistrictsForState */
export const listCandidateCitiesForState = listCandidateDistrictsForState;

/** @deprecated Use normalizeCandidateDistrict */
export const normalizeCandidateCity = normalizeCandidateDistrict;

/** @deprecated Use isKnownCandidateDistrictForState */
export const isKnownCandidateCityForState = isKnownCandidateDistrictForState;

/** @deprecated Use candidateDistrictError */
export const candidateCityError = candidateDistrictError;

/** @deprecated Use CANDIDATE_STATE_DISTRICT_MAP */
export const CANDIDATE_STATE_CITY_MAP = CANDIDATE_STATE_DISTRICT_MAP;
"""


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--json-output", type=Path, default=DEFAULT_JSON_OUTPUT)
    args = parser.parse_args()

    state_cities = extract_state_city_map(args.source)
    args.output.write_text(render_typescript(state_cities), encoding="utf-8")
    args.json_output.write_text(json.dumps(state_cities, indent=2), encoding="utf-8")
    print(f"Wrote {args.output} and {args.json_output} with {len(state_cities)} states")


if __name__ == "__main__":
    main()
